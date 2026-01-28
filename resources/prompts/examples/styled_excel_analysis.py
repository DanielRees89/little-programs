"""
STYLED EXCEL ANALYSIS TOOL
==========================
Creates professional Excel reports with:
- Maroon background with beige font throughout
- Auto-formatted headers
- Line chart AND bar chart on EACH sheet
- Commentary and insights on each sheet
- Clear summary sheet at the start

Uses openpyxl for Excel file creation.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

# ============================================
# COLOR SCHEME: Maroon & Beige
# ============================================
MAROON = '800020'          # Deep maroon background
BEIGE = 'F5F5DC'           # Beige/cream for text
LIGHT_MAROON = 'A52A2A'    # Lighter maroon for alternating rows
DARK_BEIGE = 'D4C4A8'      # Darker beige for accents

# ============================================
# STYLE DEFINITIONS
# ============================================
# Main header style (maroon bg, beige text)
header_fill = PatternFill(start_color=MAROON, end_color=MAROON, fill_type='solid')
header_font = Font(bold=True, color=BEIGE, size=12)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data cell style (maroon bg, beige text)
data_fill = PatternFill(start_color=MAROON, end_color=MAROON, fill_type='solid')
data_font = Font(color=BEIGE, size=10)
data_font_bold = Font(bold=True, color=BEIGE, size=10)

# Alternating row (slightly lighter maroon)
alt_fill = PatternFill(start_color=LIGHT_MAROON, end_color=LIGHT_MAROON, fill_type='solid')

# Title style
title_fill = PatternFill(start_color=MAROON, end_color=MAROON, fill_type='solid')
title_font = Font(bold=True, color=BEIGE, size=16)

# Commentary style
comment_fill = PatternFill(start_color='2F1515', end_color='2F1515', fill_type='solid')
comment_font = Font(italic=True, color=BEIGE, size=11)

# Border
beige_border = Border(
    left=Side(style='thin', color=DARK_BEIGE),
    right=Side(style='thin', color=DARK_BEIGE),
    top=Side(style='thin', color=DARK_BEIGE),
    bottom=Side(style='thin', color=DARK_BEIGE)
)

# Number formats
currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
percent_format = '0.0%'
number_format = '#,##0'

def style_cell(cell, is_header=False, is_alt_row=False, is_number=False, fmt=None):
    """Apply consistent styling to a cell"""
    if is_header:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    else:
        cell.fill = alt_fill if is_alt_row else data_fill
        cell.font = data_font
        cell.alignment = Alignment(horizontal='right' if is_number else 'left', vertical='center')

    cell.border = beige_border
    if fmt:
        cell.number_format = fmt

def add_title(ws, title, row=1, cols=6):
    """Add a styled title row"""
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).fill = title_fill
    ws.cell(row=row, column=1).font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

def add_commentary(ws, text, start_row, cols=6):
    """Add commentary section with styled background"""
    ws.cell(row=start_row, column=1, value=text)
    ws.cell(row=start_row, column=1).fill = comment_fill
    ws.cell(row=start_row, column=1).font = comment_font
    ws.cell(row=start_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=cols)
    ws.row_dimensions[start_row].height = 60

def auto_fit_columns(ws, min_width=12, max_width=30):
    """Auto-fit column widths based on content"""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_styled_bar_chart(title, y_title, data_ref, cats_ref, width=15, height=10):
    """Create a styled bar chart"""
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.width = width
    chart.height = height
    # Style the bars with maroon color
    for series in chart.series:
        series.graphicalProperties.solidFill = MAROON
    return chart

def create_styled_line_chart(title, y_title, data_ref, cats_ref, width=15, height=10):
    """Create a styled line chart"""
    chart = LineChart()
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.width = width
    chart.height = height
    # Style the line with maroon color
    for series in chart.series:
        series.graphicalProperties.line.solidFill = MAROON
        series.graphicalProperties.line.width = 25000  # 2.5pt
    return chart

# ============================================
# PREPARE DATA (df is pre-loaded from CSV)
# ============================================
print("Analyzing your data...")

# Get basic stats
total_rows = len(df)
columns = df.columns.tolist()

# Detect numeric columns for analysis
numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
categorical_cols = df.select_dtypes(include=['object']).columns.tolist()

print(f"Found {len(numeric_cols)} numeric columns and {len(categorical_cols)} categorical columns")

# Create summary statistics
summary_stats = df[numeric_cols].describe().T
summary_stats['sum'] = df[numeric_cols].sum()

# Calculate key insights
insights = []
for col in numeric_cols[:5]:  # Top 5 numeric columns
    insights.append({
        'metric': col,
        'total': df[col].sum(),
        'average': df[col].mean(),
        'max': df[col].max(),
        'min': df[col].min()
    })

# ============================================
# CREATE WORKBOOK
# ============================================
wb = Workbook()

# ============================================
# SHEET 1: SUMMARY (Overview Dashboard)
# ============================================
ws_summary = wb.active
ws_summary.title = "Summary"

# Apply maroon background to entire visible area
for row in range(1, 40):
    for col in range(1, 10):
        ws_summary.cell(row=row, column=col).fill = data_fill

# Title
add_title(ws_summary, 'DATA ANALYSIS SUMMARY REPORT', row=1, cols=8)
ws_summary.cell(row=2, column=1, value=f'Generated: {datetime.now().strftime("%B %d, %Y at %H:%M")}')
ws_summary.cell(row=2, column=1).font = Font(italic=True, color=BEIGE, size=10)
ws_summary.cell(row=2, column=1).fill = data_fill

# Key Statistics Section
ws_summary.cell(row=4, column=1, value='KEY STATISTICS')
ws_summary.cell(row=4, column=1).font = data_font_bold
ws_summary.cell(row=4, column=1).fill = data_fill

stats_data = [
    ('Total Records', total_rows),
    ('Numeric Columns', len(numeric_cols)),
    ('Categorical Columns', len(categorical_cols)),
    ('Total Columns', len(columns)),
]

for i, (label, value) in enumerate(stats_data, start=5):
    ws_summary.cell(row=i, column=1, value=label).font = data_font
    ws_summary.cell(row=i, column=1).fill = data_fill
    ws_summary.cell(row=i, column=2, value=value).font = data_font_bold
    ws_summary.cell(row=i, column=2).fill = data_fill
    ws_summary.cell(row=i, column=2).number_format = number_format

# Top Metrics Summary
ws_summary.cell(row=10, column=1, value='TOP METRICS OVERVIEW')
ws_summary.cell(row=10, column=1).font = data_font_bold
ws_summary.cell(row=10, column=1).fill = data_fill

# Headers for metrics table
metric_headers = ['Metric', 'Total', 'Average', 'Maximum', 'Minimum']
for col, header in enumerate(metric_headers, start=1):
    style_cell(ws_summary.cell(row=11, column=col, value=header), is_header=True)

# Metric data
for row_idx, insight in enumerate(insights, start=12):
    style_cell(ws_summary.cell(row=row_idx, column=1, value=insight['metric']), is_alt_row=row_idx%2==0)
    style_cell(ws_summary.cell(row=row_idx, column=2, value=insight['total']), is_alt_row=row_idx%2==0, is_number=True, fmt=number_format)
    style_cell(ws_summary.cell(row=row_idx, column=3, value=insight['average']), is_alt_row=row_idx%2==0, is_number=True, fmt='#,##0.00')
    style_cell(ws_summary.cell(row=row_idx, column=4, value=insight['max']), is_alt_row=row_idx%2==0, is_number=True, fmt=number_format)
    style_cell(ws_summary.cell(row=row_idx, column=5, value=insight['min']), is_alt_row=row_idx%2==0, is_number=True, fmt=number_format)

# Charts for Summary (using first numeric column)
chart_data_row = 20
if len(insights) > 0:
    # Prepare chart data
    for i, insight in enumerate(insights, start=chart_data_row):
        ws_summary.cell(row=i, column=7, value=insight['metric'][:15])
        ws_summary.cell(row=i, column=8, value=insight['total'])

    # Bar Chart
    data_ref = Reference(ws_summary, min_col=8, min_row=chart_data_row, max_row=chart_data_row + len(insights) - 1)
    cats_ref = Reference(ws_summary, min_col=7, min_row=chart_data_row, max_row=chart_data_row + len(insights) - 1)
    bar_chart = create_styled_bar_chart('Totals by Metric', 'Value', data_ref, cats_ref, width=12, height=8)
    ws_summary.add_chart(bar_chart, "A19")

    # Line Chart
    line_chart = create_styled_line_chart('Metric Trend', 'Value', data_ref, cats_ref, width=12, height=8)
    ws_summary.add_chart(line_chart, "A35")

# Commentary
add_commentary(ws_summary,
    f"EXECUTIVE SUMMARY: This dataset contains {total_rows:,} records across {len(columns)} columns. "
    f"The analysis covers {len(numeric_cols)} numeric metrics and {len(categorical_cols)} categorical dimensions. "
    f"Key insights and detailed breakdowns are provided in the following sheets.",
    start_row=52, cols=8)

# Hide chart data columns
ws_summary.column_dimensions['G'].hidden = True
ws_summary.column_dimensions['H'].hidden = True

auto_fit_columns(ws_summary)
print("Created: Summary sheet")

# ============================================
# SHEET 2: DETAILED DATA
# ============================================
ws_data = wb.create_sheet("Detailed Data")

# Apply maroon background
for row in range(1, min(total_rows + 20, 100)):
    for col in range(1, len(columns) + 3):
        ws_data.cell(row=row, column=col).fill = data_fill

add_title(ws_data, 'DETAILED DATA VIEW', row=1, cols=min(len(columns), 10))

# Headers
for col, header in enumerate(columns, start=1):
    style_cell(ws_data.cell(row=3, column=col, value=header), is_header=True)

# Data rows (limit to first 100 for performance)
display_rows = min(total_rows, 100)
for row_idx, (_, row) in enumerate(df.head(display_rows).iterrows(), start=4):
    for col_idx, col_name in enumerate(columns, start=1):
        value = row[col_name]
        cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
        is_numeric = col_name in numeric_cols
        fmt = number_format if is_numeric and pd.notna(value) else None
        style_cell(cell, is_alt_row=row_idx%2==0, is_number=is_numeric, fmt=fmt)

# Charts - use first two numeric columns if available
chart_col = len(columns) + 2
if len(numeric_cols) >= 1:
    # Prepare chart data
    chart_start = 3
    for i in range(min(20, display_rows)):
        ws_data.cell(row=chart_start + i, column=chart_col, value=i + 1)  # Row number as category
        ws_data.cell(row=chart_start + i, column=chart_col + 1, value=df[numeric_cols[0]].iloc[i] if i < len(df) else 0)

    data_ref = Reference(ws_data, min_col=chart_col + 1, min_row=chart_start, max_row=chart_start + min(20, display_rows) - 1)
    cats_ref = Reference(ws_data, min_col=chart_col, min_row=chart_start, max_row=chart_start + min(20, display_rows) - 1)

    bar_chart = create_styled_bar_chart(f'{numeric_cols[0]} Distribution', 'Value', data_ref, cats_ref)
    ws_data.add_chart(bar_chart, f"{chr(65 + len(columns) + 3)}3")

    line_chart = create_styled_line_chart(f'{numeric_cols[0]} Trend', 'Value', data_ref, cats_ref)
    ws_data.add_chart(line_chart, f"{chr(65 + len(columns) + 3)}20")

# Hide chart data columns
ws_data.column_dimensions[chr(65 + chart_col - 1)].hidden = True
ws_data.column_dimensions[chr(65 + chart_col)].hidden = True

# Commentary
comment_row = display_rows + 6
add_commentary(ws_data,
    f"DATA OVERVIEW: Displaying {display_rows:,} of {total_rows:,} total records. "
    f"The bar chart shows the distribution of {numeric_cols[0] if numeric_cols else 'values'} across the first 20 rows, "
    f"while the line chart reveals the trend pattern. Notable variations indicate areas worth investigating.",
    start_row=comment_row, cols=min(len(columns), 10))

ws_data.freeze_panes = 'A4'
auto_fit_columns(ws_data)
print("Created: Detailed Data sheet")

# ============================================
# SHEET 3: STATISTICAL ANALYSIS
# ============================================
ws_stats = wb.create_sheet("Statistical Analysis")

# Apply maroon background
for row in range(1, 50):
    for col in range(1, 12):
        ws_stats.cell(row=row, column=col).fill = data_fill

add_title(ws_stats, 'STATISTICAL ANALYSIS', row=1, cols=8)

# Summary statistics table
stat_headers = ['Metric', 'Count', 'Mean', 'Std Dev', 'Min', '25%', '50%', '75%', 'Max', 'Sum']
for col, header in enumerate(stat_headers, start=1):
    style_cell(ws_stats.cell(row=3, column=col, value=header), is_header=True)

for row_idx, col_name in enumerate(numeric_cols[:10], start=4):  # Limit to 10 metrics
    stats = df[col_name].describe()
    style_cell(ws_stats.cell(row=row_idx, column=1, value=col_name), is_alt_row=row_idx%2==0)
    style_cell(ws_stats.cell(row=row_idx, column=2, value=stats['count']), is_alt_row=row_idx%2==0, is_number=True, fmt=number_format)
    style_cell(ws_stats.cell(row=row_idx, column=3, value=stats['mean']), is_alt_row=row_idx%2==0, is_number=True, fmt='#,##0.00')
    style_cell(ws_stats.cell(row=row_idx, column=4, value=stats['std']), is_alt_row=row_idx%2==0, is_number=True, fmt='#,##0.00')
    style_cell(ws_stats.cell(row=row_idx, column=5, value=stats['min']), is_alt_row=row_idx%2==0, is_number=True, fmt=number_format)
    style_cell(ws_stats.cell(row=row_idx, column=6, value=stats['25%']), is_alt_row=row_idx%2==0, is_number=True, fmt='#,##0.00')
    style_cell(ws_stats.cell(row=row_idx, column=7, value=stats['50%']), is_alt_row=row_idx%2==0, is_number=True, fmt='#,##0.00')
    style_cell(ws_stats.cell(row=row_idx, column=8, value=stats['75%']), is_alt_row=row_idx%2==0, is_number=True, fmt='#,##0.00')
    style_cell(ws_stats.cell(row=row_idx, column=9, value=stats['max']), is_alt_row=row_idx%2==0, is_number=True, fmt=number_format)
    style_cell(ws_stats.cell(row=row_idx, column=10, value=df[col_name].sum()), is_alt_row=row_idx%2==0, is_number=True, fmt=number_format)

# Charts for statistics
chart_data_row = 20
if len(numeric_cols) >= 1:
    # Mean comparison chart data
    for i, col_name in enumerate(numeric_cols[:8], start=chart_data_row):
        ws_stats.cell(row=i, column=11, value=col_name[:12])
        ws_stats.cell(row=i, column=12, value=df[col_name].mean())

    data_ref = Reference(ws_stats, min_col=12, min_row=chart_data_row, max_row=chart_data_row + min(8, len(numeric_cols)) - 1)
    cats_ref = Reference(ws_stats, min_col=11, min_row=chart_data_row, max_row=chart_data_row + min(8, len(numeric_cols)) - 1)

    bar_chart = create_styled_bar_chart('Mean Values Comparison', 'Mean', data_ref, cats_ref)
    ws_stats.add_chart(bar_chart, "A16")

    line_chart = create_styled_line_chart('Mean Values Trend', 'Mean', data_ref, cats_ref)
    ws_stats.add_chart(line_chart, "A33")

# Hide chart data
ws_stats.column_dimensions['K'].hidden = True
ws_stats.column_dimensions['L'].hidden = True

# Commentary
top_metric = numeric_cols[0] if numeric_cols else 'N/A'
top_value = df[numeric_cols[0]].sum() if numeric_cols else 0
add_commentary(ws_stats,
    f"STATISTICAL INSIGHTS: The analysis reveals that '{top_metric}' has the highest activity with a total of {top_value:,.0f}. "
    f"The standard deviation values indicate the spread of data - higher values suggest more variability. "
    f"Consider focusing on metrics with high means but low standard deviations for consistent performance indicators.",
    start_row=50, cols=10)

auto_fit_columns(ws_stats)
print("Created: Statistical Analysis sheet")

# ============================================
# SHEET 4: PERFORMANCE BREAKDOWN
# ============================================
ws_perf = wb.create_sheet("Performance Breakdown")

# Apply maroon background
for row in range(1, 50):
    for col in range(1, 10):
        ws_perf.cell(row=row, column=col).fill = data_fill

add_title(ws_perf, 'PERFORMANCE BREAKDOWN', row=1, cols=8)

# Top performers section
ws_perf.cell(row=3, column=1, value='TOP PERFORMERS')
ws_perf.cell(row=3, column=1).font = data_font_bold
ws_perf.cell(row=3, column=1).fill = data_fill

if numeric_cols:
    main_metric = numeric_cols[0]
    top_n = df.nlargest(10, main_metric) if main_metric in df.columns else df.head(10)

    # Headers
    display_cols = columns[:6]  # Show first 6 columns
    for col, header in enumerate(display_cols, start=1):
        style_cell(ws_perf.cell(row=4, column=col, value=header), is_header=True)

    # Data
    for row_idx, (_, row) in enumerate(top_n.iterrows(), start=5):
        for col_idx, col_name in enumerate(display_cols, start=1):
            value = row[col_name]
            cell = ws_perf.cell(row=row_idx, column=col_idx, value=value)
            is_numeric = col_name in numeric_cols
            style_cell(cell, is_alt_row=row_idx%2==0, is_number=is_numeric)

# Charts
chart_data_row = 20
for i, (_, row) in enumerate(top_n.head(8).iterrows(), start=chart_data_row):
    label = str(row[columns[0]])[:15] if len(columns) > 0 else f"Item {i}"
    ws_perf.cell(row=i, column=7, value=label)
    ws_perf.cell(row=i, column=8, value=row[main_metric] if main_metric in row else 0)

data_ref = Reference(ws_perf, min_col=8, min_row=chart_data_row, max_row=chart_data_row + min(8, len(top_n)) - 1)
cats_ref = Reference(ws_perf, min_col=7, min_row=chart_data_row, max_row=chart_data_row + min(8, len(top_n)) - 1)

bar_chart = create_styled_bar_chart('Top Performers', main_metric, data_ref, cats_ref)
ws_perf.add_chart(bar_chart, "A17")

line_chart = create_styled_line_chart('Performance Trend', main_metric, data_ref, cats_ref)
ws_perf.add_chart(line_chart, "A34")

# Hide chart data
ws_perf.column_dimensions['G'].hidden = True
ws_perf.column_dimensions['H'].hidden = True

# Commentary
top_performer = top_n.iloc[0][columns[0]] if len(top_n) > 0 and len(columns) > 0 else 'N/A'
top_value = top_n.iloc[0][main_metric] if len(top_n) > 0 else 0
add_commentary(ws_perf,
    f"PERFORMANCE ANALYSIS: The top performer is '{top_performer}' with {top_value:,.0f} in {main_metric}. "
    f"The bar chart visualizes the ranking of top 8 performers, while the line chart shows the performance curve. "
    f"A steep decline in the line chart indicates a small group of high performers dominating the results.",
    start_row=51, cols=8)

auto_fit_columns(ws_perf)
print("Created: Performance Breakdown sheet")

# ============================================
# SAVE WORKBOOK
# ============================================
filename = 'styled_analysis_report.xlsx'
wb.save(filename)

print(f"\n{'='*60}")
print(f"EXCEL REPORT GENERATED SUCCESSFULLY!")
print(f"{'='*60}")
print(f"\nFilename: {filename}")
print(f"\nSheets created:")
print(f"  1. Summary - Executive overview with key metrics")
print(f"  2. Detailed Data - Full data view with formatting")
print(f"  3. Statistical Analysis - Comprehensive statistics")
print(f"  4. Performance Breakdown - Top performers analysis")
print(f"\nFeatures:")
print(f"  - Maroon background with beige text throughout")
print(f"  - Auto-formatted headers and columns")
print(f"  - Bar chart AND line chart on each sheet")
print(f"  - Commentary and insights on every sheet")
print(f"{'='*60}")
