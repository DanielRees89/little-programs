"""
SALES ANALYSIS EXCEL REPORT
============================
Creates a professional Excel report from sales SKU data with:
- Maroon background (#800000) with beige font (#F5F5DC) throughout
- Auto-formatted headers with proper column widths
- Line chart AND bar chart on EACH sheet
- Commentary and insights on each sheet
- Clear summary sheet at the start

Sheets:
1. Summary - Key metrics overview
2. 7-Day Performance - Recent week analysis
3. 30-Day Performance - Monthly analysis
4. 90-Day Performance - Quarterly analysis
5. Period Comparison - Side-by-side trend analysis

Usage: The CSV data should be pre-loaded as 'df' DataFrame
Expected columns: SKU, Product Name, Category, Color,
                  7D Units, 7D Sales, 7D Net Sales, 7D Avg Price,
                  30D Units, 30D Sales, 30D Net Sales, 30D Avg Price,
                  90D Units, 90D Sales, 90D Net Sales, 90D Avg Price
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# ============================================
# COLOR SCHEME: Maroon & Beige
# ============================================
MAROON = '800000'           # Deep maroon background
BEIGE = 'F5F5DC'            # Beige/cream for text
LIGHT_MAROON = 'A52A2A'     # Lighter maroon for alternating rows
DARK_MAROON = '4A0000'      # Darker maroon for headers
DARK_BEIGE = 'D4C4A8'       # Darker beige for accents
COMMENT_BG = '2F1515'       # Dark background for commentary

# ============================================
# STYLE DEFINITIONS
# ============================================
header_fill = PatternFill(start_color=DARK_MAROON, end_color=DARK_MAROON, fill_type='solid')
header_font = Font(bold=True, color=BEIGE, size=12)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_fill = PatternFill(start_color=MAROON, end_color=MAROON, fill_type='solid')
data_font = Font(color=BEIGE, size=10)
data_font_bold = Font(bold=True, color=BEIGE, size=10)

alt_fill = PatternFill(start_color=LIGHT_MAROON, end_color=LIGHT_MAROON, fill_type='solid')

title_fill = PatternFill(start_color=MAROON, end_color=MAROON, fill_type='solid')
title_font = Font(bold=True, color=BEIGE, size=18)

subtitle_font = Font(bold=True, color=BEIGE, size=14)

comment_fill = PatternFill(start_color=COMMENT_BG, end_color=COMMENT_BG, fill_type='solid')
comment_font = Font(italic=True, color=BEIGE, size=11)

beige_border = Border(
    left=Side(style='thin', color=DARK_BEIGE),
    right=Side(style='thin', color=DARK_BEIGE),
    top=Side(style='thin', color=DARK_BEIGE),
    bottom=Side(style='thin', color=DARK_BEIGE)
)

# Number formats
currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
number_format = '#,##0'
decimal_format = '#,##0.00'


def style_cell(cell, is_header=False, is_alt_row=False, is_number=False, fmt=None):
    """Apply consistent styling to a cell"""
    if is_header:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    else:
        cell.fill = alt_fill if is_alt_row else data_fill
        cell.font = data_font
        cell.alignment = Alignment(horizontal='right' if is_number else 'left', vertical='center')

    cell.border = beige_border
    if fmt:
        cell.number_format = fmt


def fill_background(ws, rows=60, cols=15):
    """Fill entire visible area with maroon background"""
    for row in range(1, rows):
        for col in range(1, cols):
            ws.cell(row=row, column=col).fill = data_fill


def add_title(ws, title, row=1, cols=8):
    """Add a styled title row"""
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = title_fill
    cell.font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 30


def add_subtitle(ws, text, row, cols=8):
    """Add a subtitle row"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = data_fill
    cell.font = subtitle_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)


def add_commentary(ws, text, start_row, cols=8):
    """Add commentary section with styled background"""
    cell = ws.cell(row=start_row, column=1, value=text)
    cell.fill = comment_fill
    cell.font = comment_font
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 1, end_column=cols)
    ws.row_dimensions[start_row].height = 45
    ws.row_dimensions[start_row + 1].height = 30


def auto_fit_columns(ws, min_width=10, max_width=25):
    """Auto-fit column widths based on content"""
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and not isinstance(cell, type(None)):
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_bar_chart(title, y_title, data_ref, cats_ref, width=14, height=10):
    """Create a styled bar chart"""
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.width = width
    chart.height = height
    chart.legend = None
    # Style bars with maroon color
    for series in chart.series:
        series.graphicalProperties.solidFill = BEIGE
    return chart


def create_line_chart(title, y_title, data_refs, cats_ref, series_names=None, width=14, height=10):
    """Create a styled line chart with multiple series"""
    chart = LineChart()
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title

    colors = [BEIGE, 'FFD700', 'CD853F']  # Beige, Gold, Peru

    for i, data_ref in enumerate(data_refs):
        chart.add_data(data_ref, titles_from_data=False)

    chart.set_categories(cats_ref)
    chart.width = width
    chart.height = height

    # Style lines
    for i, series in enumerate(chart.series):
        series.graphicalProperties.line.solidFill = colors[i % len(colors)]
        series.graphicalProperties.line.width = 25000  # 2.5pt
        if series_names and i < len(series_names):
            series.tx = SeriesLabel(v=series_names[i])

    return chart


def write_data_table(ws, df, start_row, columns=None):
    """Write a DataFrame to worksheet with styling"""
    if columns is None:
        columns = df.columns.tolist()

    # Headers
    for col_idx, col_name in enumerate(columns, start=1):
        style_cell(ws.cell(row=start_row, column=col_idx, value=col_name), is_header=True)

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)

            # Handle numeric formatting
            if pd.isna(value):
                cell.value = ''
            elif isinstance(value, (int, float)):
                cell.value = value
                if 'Price' in col_name or 'Sales' in col_name:
                    fmt = currency_format
                elif 'Units' in col_name:
                    fmt = number_format
                else:
                    fmt = decimal_format
                style_cell(cell, is_alt_row=(row_idx % 2 == 0), is_number=True, fmt=fmt)
            else:
                cell.value = value
                style_cell(cell, is_alt_row=(row_idx % 2 == 0))

    return start_row + len(df) + 1


# ============================================
# MAIN REPORT GENERATION
# ============================================
print("Analyzing sales data...")

# Extract size from SKU for cleaner analysis
df['Size'] = df['SKU'].str.extract(r'-(\d+[RSL]?)$')[0].fillna(df['SKU'].str[-3:])

# Calculate key metrics
total_7d_units = df['7D Units'].sum()
total_7d_sales = df['7D Sales'].sum()
total_7d_net = df['7D Net Sales'].sum()
total_30d_units = df['30D Units'].sum()
total_30d_sales = df['30D Sales'].sum()
total_30d_net = df['30D Net Sales'].sum()
total_90d_units = df['90D Units'].sum()
total_90d_sales = df['90D Sales'].sum()
total_90d_net = df['90D Net Sales'].sum()

avg_7d_price = df['7D Avg Price'].mean()
avg_30d_price = df['30D Avg Price'].mean()
avg_90d_price = df['90D Avg Price'].mean()

top_seller_7d = df.loc[df['7D Units'].idxmax(), 'Size'] if total_7d_units > 0 else 'N/A'
top_seller_30d = df.loc[df['30D Units'].idxmax(), 'Size'] if total_30d_units > 0 else 'N/A'
top_seller_90d = df.loc[df['90D Units'].idxmax(), 'Size'] if total_90d_units > 0 else 'N/A'

product_name = df['Product Name'].iloc[0] if 'Product Name' in df.columns else 'Product'
category = df['Category'].iloc[0] if 'Category' in df.columns else 'Category'
color = df['Color'].iloc[0] if 'Color' in df.columns else 'Color'

# Create workbook
wb = Workbook()

# ============================================
# SHEET 1: SUMMARY
# ============================================
ws_summary = wb.active
ws_summary.title = "Summary"

fill_background(ws_summary, rows=55, cols=12)

add_title(ws_summary, 'SALES ANALYSIS SUMMARY', row=1, cols=8)

# Subtitle with date
ws_summary.cell(row=2, column=1, value=f'Generated: {datetime.now().strftime("%B %d, %Y")}')
ws_summary.cell(row=2, column=1).font = Font(italic=True, color=BEIGE, size=10)

# Product info
add_subtitle(ws_summary, f'{category} - {color}', row=4, cols=8)

# Key metrics table
metrics_row = 6
ws_summary.cell(row=metrics_row, column=1, value='KEY METRICS').font = data_font_bold

headers = ['Period', 'Units Sold', 'Gross Sales', 'Net Sales', 'Avg Price', 'Top Size']
for col, header in enumerate(headers, start=1):
    style_cell(ws_summary.cell(row=metrics_row + 1, column=col, value=header), is_header=True)

# 7 Day row
metrics_data = [
    ['7 Days', total_7d_units, total_7d_sales, total_7d_net, avg_7d_price, top_seller_7d],
    ['30 Days', total_30d_units, total_30d_sales, total_30d_net, avg_30d_price, top_seller_30d],
    ['90 Days', total_90d_units, total_90d_sales, total_90d_net, avg_90d_price, top_seller_90d],
]

for row_idx, row_data in enumerate(metrics_data, start=metrics_row + 2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        is_num = col_idx in [2, 3, 4, 5]
        fmt = currency_format if col_idx in [3, 4, 5] else (number_format if col_idx == 2 else None)
        style_cell(cell, is_alt_row=(row_idx % 2 == 0), is_number=is_num, fmt=fmt)

# Chart data (hidden columns)
chart_row = 15
periods = ['7 Days', '30 Days', '90 Days']
units = [total_7d_units, total_30d_units, total_90d_units]
sales = [total_7d_sales, total_30d_sales, total_90d_sales]
net_sales = [total_7d_net, total_30d_net, total_90d_net]

for i, (period, unit, sale, net) in enumerate(zip(periods, units, sales, net_sales)):
    ws_summary.cell(row=chart_row + i, column=10, value=period)
    ws_summary.cell(row=chart_row + i, column=11, value=unit)
    ws_summary.cell(row=chart_row + i, column=12, value=sale)
    ws_summary.cell(row=chart_row + i, column=13, value=net)

# Bar Chart - Units by Period
cats_ref = Reference(ws_summary, min_col=10, min_row=chart_row, max_row=chart_row + 2)
data_ref = Reference(ws_summary, min_col=11, min_row=chart_row, max_row=chart_row + 2)
bar_chart = create_bar_chart('Total Units by Period', 'Units', data_ref, cats_ref)
ws_summary.add_chart(bar_chart, 'A13')

# Line Chart - Sales Trend
data_ref1 = Reference(ws_summary, min_col=12, min_row=chart_row, max_row=chart_row + 2)
data_ref2 = Reference(ws_summary, min_col=13, min_row=chart_row, max_row=chart_row + 2)
line_chart = create_line_chart('Sales Trend by Period', 'Sales ($)', [data_ref1, data_ref2], cats_ref,
                                series_names=['Gross Sales', 'Net Sales'])
ws_summary.add_chart(line_chart, 'A28')

# Hide chart data columns
for col in ['J', 'K', 'L', 'M']:
    ws_summary.column_dimensions[col].hidden = True

# Commentary
growth_7_to_30 = ((total_30d_units / total_7d_units * 30/7) - 1) * 100 if total_7d_units > 0 else 0
add_commentary(ws_summary,
    f"EXECUTIVE SUMMARY: This report analyzes {len(df)} SKUs for {category} in {color}. "
    f"Over the past 90 days, total sales reached ${total_90d_sales:,.2f} from {total_90d_units:,} units. "
    f"Size {top_seller_90d} is the best performer. The 7-day run rate suggests "
    f"{'strong momentum' if total_7d_units > total_30d_units/4 else 'steady performance'}. "
    f"See individual period sheets for detailed breakdowns.",
    start_row=44, cols=8)

auto_fit_columns(ws_summary)
print("  Created: Summary sheet")

# ============================================
# SHEET 2: 7-DAY PERFORMANCE
# ============================================
ws_7d = wb.create_sheet("7-Day Performance")

fill_background(ws_7d, rows=60, cols=12)
add_title(ws_7d, '7-DAY PERFORMANCE ANALYSIS', row=1, cols=7)

# Prepare data sorted by units
df_7d = df[['Size', '7D Units', '7D Sales', '7D Net Sales', '7D Avg Price']].copy()
df_7d = df_7d.sort_values('7D Units', ascending=False)
df_7d.columns = ['Size', 'Units', 'Gross Sales', 'Net Sales', 'Avg Price']

# Write table
table_end = write_data_table(ws_7d, df_7d, start_row=3)

# Chart data
chart_row = 3
top_10 = df_7d.head(10)
for i, (_, row) in enumerate(top_10.iterrows()):
    ws_7d.cell(row=chart_row + i, column=9, value=row['Size'])
    ws_7d.cell(row=chart_row + i, column=10, value=row['Units'])
    ws_7d.cell(row=chart_row + i, column=11, value=row['Gross Sales'])
    ws_7d.cell(row=chart_row + i, column=12, value=row['Net Sales'])

# Bar Chart
cats_ref = Reference(ws_7d, min_col=9, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
data_ref = Reference(ws_7d, min_col=10, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
bar_chart = create_bar_chart('Units Sold by Size (Top 10)', 'Units', data_ref, cats_ref)
ws_7d.add_chart(bar_chart, 'A' + str(table_end + 2))

# Line Chart
data_ref1 = Reference(ws_7d, min_col=11, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
data_ref2 = Reference(ws_7d, min_col=12, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
line_chart = create_line_chart('Gross vs Net Sales by Size', 'Sales ($)', [data_ref1, data_ref2], cats_ref,
                                series_names=['Gross Sales', 'Net Sales'])
ws_7d.add_chart(line_chart, 'A' + str(table_end + 15))

# Hide chart data
for col in ['I', 'J', 'K', 'L']:
    ws_7d.column_dimensions[col].hidden = True

# Commentary
best_size_7d = df_7d.iloc[0]['Size'] if len(df_7d) > 0 else 'N/A'
best_units_7d = df_7d.iloc[0]['Units'] if len(df_7d) > 0 else 0
add_commentary(ws_7d,
    f"7-DAY INSIGHTS: Size {best_size_7d} leads with {best_units_7d:.0f} units sold this week. "
    f"Total weekly revenue is ${total_7d_sales:,.2f} with net sales of ${total_7d_net:,.2f}. "
    f"The bar chart shows the distribution of sales across sizes, while the line chart "
    f"highlights the margin between gross and net sales. Monitor sizes with low movement for potential markdown opportunities.",
    start_row=table_end + 28, cols=7)

auto_fit_columns(ws_7d)
print("  Created: 7-Day Performance sheet")

# ============================================
# SHEET 3: 30-DAY PERFORMANCE
# ============================================
ws_30d = wb.create_sheet("30-Day Performance")

fill_background(ws_30d, rows=60, cols=12)
add_title(ws_30d, '30-DAY PERFORMANCE ANALYSIS', row=1, cols=7)

df_30d = df[['Size', '30D Units', '30D Sales', '30D Net Sales', '30D Avg Price']].copy()
df_30d = df_30d.sort_values('30D Units', ascending=False)
df_30d.columns = ['Size', 'Units', 'Gross Sales', 'Net Sales', 'Avg Price']

table_end = write_data_table(ws_30d, df_30d, start_row=3)

# Chart data
chart_row = 3
top_10 = df_30d.head(10)
for i, (_, row) in enumerate(top_10.iterrows()):
    ws_30d.cell(row=chart_row + i, column=9, value=row['Size'])
    ws_30d.cell(row=chart_row + i, column=10, value=row['Units'])
    ws_30d.cell(row=chart_row + i, column=11, value=row['Gross Sales'])
    ws_30d.cell(row=chart_row + i, column=12, value=row['Net Sales'])

# Bar Chart
cats_ref = Reference(ws_30d, min_col=9, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
data_ref = Reference(ws_30d, min_col=10, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
bar_chart = create_bar_chart('Units Sold by Size (Top 10)', 'Units', data_ref, cats_ref)
ws_30d.add_chart(bar_chart, 'A' + str(table_end + 2))

# Line Chart
data_ref1 = Reference(ws_30d, min_col=11, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
data_ref2 = Reference(ws_30d, min_col=12, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
line_chart = create_line_chart('Gross vs Net Sales by Size', 'Sales ($)', [data_ref1, data_ref2], cats_ref,
                                series_names=['Gross Sales', 'Net Sales'])
ws_30d.add_chart(line_chart, 'A' + str(table_end + 15))

for col in ['I', 'J', 'K', 'L']:
    ws_30d.column_dimensions[col].hidden = True

best_size_30d = df_30d.iloc[0]['Size'] if len(df_30d) > 0 else 'N/A'
best_units_30d = df_30d.iloc[0]['Units'] if len(df_30d) > 0 else 0
add_commentary(ws_30d,
    f"30-DAY INSIGHTS: Size {best_size_30d} is the monthly leader with {best_units_30d:.0f} units. "
    f"Monthly revenue reached ${total_30d_sales:,.2f}. Average selling price is ${avg_30d_price:,.2f}. "
    f"Compare the sales distribution to identify consistent performers vs. one-time spikes. "
    f"Sizes appearing in both 7D and 30D top performers indicate sustained demand.",
    start_row=table_end + 28, cols=7)

auto_fit_columns(ws_30d)
print("  Created: 30-Day Performance sheet")

# ============================================
# SHEET 4: 90-DAY PERFORMANCE
# ============================================
ws_90d = wb.create_sheet("90-Day Performance")

fill_background(ws_90d, rows=60, cols=12)
add_title(ws_90d, '90-DAY PERFORMANCE ANALYSIS', row=1, cols=7)

df_90d = df[['Size', '90D Units', '90D Sales', '90D Net Sales', '90D Avg Price']].copy()
df_90d = df_90d.sort_values('90D Units', ascending=False)
df_90d.columns = ['Size', 'Units', 'Gross Sales', 'Net Sales', 'Avg Price']

table_end = write_data_table(ws_90d, df_90d, start_row=3)

# Chart data
chart_row = 3
top_10 = df_90d.head(10)
for i, (_, row) in enumerate(top_10.iterrows()):
    ws_90d.cell(row=chart_row + i, column=9, value=row['Size'])
    ws_90d.cell(row=chart_row + i, column=10, value=row['Units'])
    ws_90d.cell(row=chart_row + i, column=11, value=row['Gross Sales'])
    ws_90d.cell(row=chart_row + i, column=12, value=row['Net Sales'])

# Bar Chart
cats_ref = Reference(ws_90d, min_col=9, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
data_ref = Reference(ws_90d, min_col=10, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
bar_chart = create_bar_chart('Units Sold by Size (Top 10)', 'Units', data_ref, cats_ref)
ws_90d.add_chart(bar_chart, 'A' + str(table_end + 2))

# Line Chart
data_ref1 = Reference(ws_90d, min_col=11, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
data_ref2 = Reference(ws_90d, min_col=12, min_row=chart_row, max_row=chart_row + min(10, len(top_10)) - 1)
line_chart = create_line_chart('Gross vs Net Sales by Size', 'Sales ($)', [data_ref1, data_ref2], cats_ref,
                                series_names=['Gross Sales', 'Net Sales'])
ws_90d.add_chart(line_chart, 'A' + str(table_end + 15))

for col in ['I', 'J', 'K', 'L']:
    ws_90d.column_dimensions[col].hidden = True

best_size_90d = df_90d.iloc[0]['Size'] if len(df_90d) > 0 else 'N/A'
best_units_90d = df_90d.iloc[0]['Units'] if len(df_90d) > 0 else 0
margin_pct = (total_90d_net / total_90d_sales * 100) if total_90d_sales > 0 else 0
add_commentary(ws_90d,
    f"90-DAY INSIGHTS: Quarterly analysis shows size {best_size_90d} as the top seller with {best_units_90d:.0f} units. "
    f"Total quarterly revenue is ${total_90d_sales:,.2f} with a net margin of {margin_pct:.1f}%. "
    f"This long-term view reveals true demand patterns and should guide inventory planning. "
    f"Consider increasing stock for top performers and reviewing slow movers.",
    start_row=table_end + 28, cols=7)

auto_fit_columns(ws_90d)
print("  Created: 90-Day Performance sheet")

# ============================================
# SHEET 5: PERIOD COMPARISON
# ============================================
ws_compare = wb.create_sheet("Period Comparison")

fill_background(ws_compare, rows=60, cols=15)
add_title(ws_compare, 'PERIOD COMPARISON ANALYSIS', row=1, cols=10)

# Comparison table
df_compare = df[['Size', '7D Units', '7D Sales', '30D Units', '30D Sales', '90D Units', '90D Sales']].copy()
df_compare = df_compare.sort_values('90D Units', ascending=False)

table_end = write_data_table(ws_compare, df_compare, start_row=3)

# Chart data - compare periods for top sizes
chart_row = 3
top_sizes = df_compare.head(8)
for i, (_, row) in enumerate(top_sizes.iterrows()):
    ws_compare.cell(row=chart_row + i, column=11, value=row['Size'])
    ws_compare.cell(row=chart_row + i, column=12, value=row['7D Units'])
    ws_compare.cell(row=chart_row + i, column=13, value=row['30D Units'])
    ws_compare.cell(row=chart_row + i, column=14, value=row['90D Units'])

# Stacked Bar Chart showing units across periods
cats_ref = Reference(ws_compare, min_col=11, min_row=chart_row, max_row=chart_row + min(8, len(top_sizes)) - 1)
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Units by Size Across Periods"
chart.y_axis.title = "Units"

data_7d = Reference(ws_compare, min_col=12, min_row=chart_row, max_row=chart_row + min(8, len(top_sizes)) - 1)
data_30d = Reference(ws_compare, min_col=13, min_row=chart_row, max_row=chart_row + min(8, len(top_sizes)) - 1)
data_90d = Reference(ws_compare, min_col=14, min_row=chart_row, max_row=chart_row + min(8, len(top_sizes)) - 1)

chart.add_data(data_7d, titles_from_data=False)
chart.add_data(data_30d, titles_from_data=False)
chart.add_data(data_90d, titles_from_data=False)
chart.set_categories(cats_ref)
chart.width = 15
chart.height = 10

colors = [BEIGE, 'FFD700', 'CD853F']
series_titles = ['7D', '30D', '90D']
for i, series in enumerate(chart.series):
    series.graphicalProperties.solidFill = colors[i]
    series.tx = SeriesLabel(v=series_titles[i])

ws_compare.add_chart(chart, 'A' + str(table_end + 2))

# Line Chart - Sales trend comparison
line_chart = LineChart()
line_chart.style = 10
line_chart.title = "Sales Trend Across Periods"
line_chart.y_axis.title = "Sales ($)"

# Use period totals for trend
period_chart_row = 35
ws_compare.cell(row=period_chart_row, column=11, value='7 Days')
ws_compare.cell(row=period_chart_row, column=12, value=total_7d_sales)
ws_compare.cell(row=period_chart_row + 1, column=11, value='30 Days')
ws_compare.cell(row=period_chart_row + 1, column=12, value=total_30d_sales)
ws_compare.cell(row=period_chart_row + 2, column=11, value='90 Days')
ws_compare.cell(row=period_chart_row + 2, column=12, value=total_90d_sales)

cats_ref2 = Reference(ws_compare, min_col=11, min_row=period_chart_row, max_row=period_chart_row + 2)
data_ref2 = Reference(ws_compare, min_col=12, min_row=period_chart_row, max_row=period_chart_row + 2)

line_chart.add_data(data_ref2, titles_from_data=False)
line_chart.set_categories(cats_ref2)
line_chart.width = 15
line_chart.height = 10

for series in line_chart.series:
    series.graphicalProperties.line.solidFill = BEIGE
    series.graphicalProperties.line.width = 30000

ws_compare.add_chart(line_chart, 'A' + str(table_end + 15))

for col in ['K', 'L', 'M', 'N']:
    ws_compare.column_dimensions[col].hidden = True

# Calculate velocity
weekly_rate = total_90d_sales / 13 if total_90d_sales > 0 else 0
recent_rate = total_7d_sales
velocity_change = ((recent_rate / weekly_rate) - 1) * 100 if weekly_rate > 0 else 0

add_commentary(ws_compare,
    f"TREND ANALYSIS: This view compares performance across all three time periods. "
    f"Weekly sales rate from 90D average: ${weekly_rate:,.2f}. Recent 7D rate: ${recent_rate:,.2f}. "
    f"This represents a {velocity_change:+.1f}% velocity change. "
    f"The bar chart shows how individual sizes perform over time, while the line chart shows overall sales trajectory. "
    f"Consistent performers across all periods are your core sellers.",
    start_row=table_end + 28, cols=10)

auto_fit_columns(ws_compare)
print("  Created: Period Comparison sheet")

# ============================================
# SAVE WORKBOOK
# ============================================
filename = 'sales_analysis_report.xlsx'
wb.save(filename)

print(f"\n{'='*60}")
print(f"EXCEL REPORT GENERATED SUCCESSFULLY!")
print(f"{'='*60}")
print(f"\nFilename: {filename}")
print(f"\nSheets created:")
print(f"  1. Summary - Executive overview with key metrics")
print(f"  2. 7-Day Performance - Weekly sales analysis")
print(f"  3. 30-Day Performance - Monthly sales analysis")
print(f"  4. 90-Day Performance - Quarterly sales analysis")
print(f"  5. Period Comparison - Cross-period trend analysis")
print(f"\nFeatures:")
print(f"  - Maroon background (#800000) with beige text (#F5F5DC)")
print(f"  - Auto-formatted headers and column widths")
print(f"  - Bar chart AND line chart on each sheet")
print(f"  - Commentary and insights on every sheet")
print(f"  - Currency and number formatting applied")
print(f"{'='*60}")
