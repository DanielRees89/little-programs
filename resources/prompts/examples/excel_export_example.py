"""
EXCEL EXPORT EXAMPLE
====================
This script shows how to create professional Excel reports with:
- Multiple sheets
- Formatted tables with headers
- Conditional formatting
- Charts embedded in Excel
- Summary dashboards

Uses openpyxl for Excel file creation.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

# ============================================
# BRAND COLORS (as hex for Excel)
# ============================================
BRAND_PRIMARY = '822F23'
BRAND_SECONDARY = 'D67D63'
BRAND_ACCENT = 'A1945F'
BRAND_LIGHT = 'F6EFDB'
BRAND_DARK = '1F2937'

# ============================================
# STYLE DEFINITIONS
# ============================================
# Header style
header_fill = PatternFill(start_color=BRAND_PRIMARY, end_color=BRAND_PRIMARY, fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data styles
data_font = Font(size=10)
data_alignment = Alignment(horizontal='left', vertical='center')
number_alignment = Alignment(horizontal='right', vertical='center')
currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
percent_format = '0.0%'
number_format = '#,##0'

# Border
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# Alternating row fill
alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

# ============================================
# PREPARE DATA (df is pre-loaded)
# ============================================
df['Day'] = pd.to_datetime(df['Day'])

# Daily summary
daily = df.groupby('Day').agg({
    'Orders': 'sum',
    'Gross sales': 'sum',
    'Discounts': 'sum',
    'Net sales': 'sum',
    'Gross profit': 'sum',
    'Quantity ordered': 'sum'
}).reset_index()

daily['AOV'] = daily['Gross sales'] / daily['Orders']
daily['Gross Margin %'] = daily['Gross profit'] / daily['Net sales']
daily['Discount Rate %'] = abs(daily['Discounts']) / daily['Gross sales']

# Customer breakdown
customer_summary = df.groupby('New or returning customer').agg({
    'Orders': 'sum',
    'Gross sales': 'sum',
    'Net sales': 'sum',
    'Gross profit': 'sum'
}).reset_index()

customer_summary['AOV'] = customer_summary['Gross sales'] / customer_summary['Orders']
customer_summary['% of Orders'] = customer_summary['Orders'] / customer_summary['Orders'].sum()
customer_summary['% of Revenue'] = customer_summary['Gross sales'] / customer_summary['Gross sales'].sum()

print(f"Data prepared: {len(daily)} days, {len(customer_summary)} customer segments")

# ============================================
# CREATE WORKBOOK
# ============================================
wb = Workbook()

# ============================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================
ws_summary = wb.active
ws_summary.title = "Executive Summary"

# Title
ws_summary['A1'] = 'BUILT DIFFERENT - November 2025 Report'
ws_summary['A1'].font = Font(bold=True, size=18, color=BRAND_PRIMARY)
ws_summary.merge_cells('A1:E1')

ws_summary['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y")}'
ws_summary['A2'].font = Font(italic=True, color='6B7280')

# KPI Summary
kpis = [
    ('Total Orders', daily['Orders'].sum(), number_format),
    ('Gross Sales', daily['Gross sales'].sum(), currency_format),
    ('Net Sales', daily['Net sales'].sum(), currency_format),
    ('Gross Profit', daily['Gross profit'].sum(), currency_format),
    ('Average Order Value', daily['Gross sales'].sum() / daily['Orders'].sum(), currency_format),
    ('Gross Margin', daily['Gross profit'].sum() / daily['Net sales'].sum(), percent_format),
    ('Total Items Sold', daily['Quantity ordered'].sum(), number_format),
]

ws_summary['A4'] = 'Key Performance Indicators'
ws_summary['A4'].font = Font(bold=True, size=14, color=BRAND_PRIMARY)

for i, (label, value, fmt) in enumerate(kpis, start=5):
    ws_summary[f'A{i}'] = label
    ws_summary[f'A{i}'].font = Font(bold=True)
    ws_summary[f'B{i}'] = value
    ws_summary[f'B{i}'].number_format = fmt
    ws_summary[f'B{i}'].alignment = number_alignment

# Column widths
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 18

print("✓ Created: Executive Summary sheet")

# ============================================
# SHEET 2: DAILY DATA
# ============================================
ws_daily = wb.create_sheet("Daily Data")

# Write headers
headers = ['Date', 'Orders', 'Gross Sales', 'Discounts', 'Net Sales', 
           'Gross Profit', 'Items Sold', 'AOV', 'Gross Margin %', 'Discount Rate %']

for col, header in enumerate(headers, start=1):
    cell = ws_daily.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
for row_idx, row in enumerate(daily.itertuples(), start=2):
    # Date
    ws_daily.cell(row=row_idx, column=1, value=row.Day.strftime('%Y-%m-%d'))
    # Orders
    ws_daily.cell(row=row_idx, column=2, value=row.Orders).number_format = number_format
    # Gross Sales
    ws_daily.cell(row=row_idx, column=3, value=row._3).number_format = currency_format
    # Discounts
    ws_daily.cell(row=row_idx, column=4, value=row.Discounts).number_format = currency_format
    # Net Sales
    ws_daily.cell(row=row_idx, column=5, value=row._5).number_format = currency_format
    # Gross Profit
    ws_daily.cell(row=row_idx, column=6, value=row._6).number_format = currency_format
    # Items Sold
    ws_daily.cell(row=row_idx, column=7, value=row._7).number_format = number_format
    # AOV
    ws_daily.cell(row=row_idx, column=8, value=row.AOV).number_format = currency_format
    # Gross Margin %
    ws_daily.cell(row=row_idx, column=9, value=row._9).number_format = percent_format
    # Discount Rate %
    ws_daily.cell(row=row_idx, column=10, value=row._10).number_format = percent_format
    
    # Apply alternating row colors and borders
    for col in range(1, 11):
        cell = ws_daily.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.font = data_font
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# Column widths
col_widths = [12, 10, 14, 14, 14, 14, 12, 12, 14, 14]
for i, width in enumerate(col_widths, start=1):
    ws_daily.column_dimensions[chr(64 + i)].width = width

# Freeze header row
ws_daily.freeze_panes = 'A2'

# Add conditional formatting for Gross Margin
ws_daily.conditional_formatting.add(
    f'I2:I{len(daily)+1}',
    ColorScaleRule(
        start_type='min', start_color='F8D7DA',
        mid_type='percentile', mid_value=50, mid_color='FFF3CD',
        end_type='max', end_color='D4EDDA'
    )
)

print("✓ Created: Daily Data sheet")

# ============================================
# SHEET 3: CUSTOMER ANALYSIS
# ============================================
ws_customer = wb.create_sheet("Customer Analysis")

# Title
ws_customer['A1'] = 'Customer Segment Analysis'
ws_customer['A1'].font = Font(bold=True, size=14, color=BRAND_PRIMARY)

# Headers
cust_headers = ['Customer Type', 'Orders', 'Gross Sales', 'Net Sales', 
                'Gross Profit', 'AOV', '% of Orders', '% of Revenue']

for col, header in enumerate(cust_headers, start=1):
    cell = ws_customer.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Data
for row_idx, row in enumerate(customer_summary.itertuples(), start=4):
    ws_customer.cell(row=row_idx, column=1, value=row._1)
    ws_customer.cell(row=row_idx, column=2, value=row.Orders).number_format = number_format
    ws_customer.cell(row=row_idx, column=3, value=row._3).number_format = currency_format
    ws_customer.cell(row=row_idx, column=4, value=row._4).number_format = currency_format
    ws_customer.cell(row=row_idx, column=5, value=row._5).number_format = currency_format
    ws_customer.cell(row=row_idx, column=6, value=row.AOV).number_format = currency_format
    ws_customer.cell(row=row_idx, column=7, value=row._7).number_format = percent_format
    ws_customer.cell(row=row_idx, column=8, value=row._8).number_format = percent_format
    
    for col in range(1, 9):
        ws_customer.cell(row=row_idx, column=col).border = thin_border

# Column widths
for i, width in enumerate([18, 12, 14, 14, 14, 12, 12, 12], start=1):
    ws_customer.column_dimensions[chr(64 + i)].width = width

print("✓ Created: Customer Analysis sheet")

# ============================================
# SHEET 4: CHARTS
# ============================================
ws_charts = wb.create_sheet("Charts")

# Copy daily data for chart references
for row_idx, row in enumerate(daily.itertuples(), start=1):
    ws_charts.cell(row=row_idx, column=1, value=row.Day.strftime('%b %d'))
    ws_charts.cell(row=row_idx, column=2, value=row.Orders)
    ws_charts.cell(row=row_idx, column=3, value=row._3)  # Gross sales

# BAR CHART - Orders
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Daily Orders"
chart1.y_axis.title = "Orders"
chart1.x_axis.title = "Date"

data1 = Reference(ws_charts, min_col=2, min_row=1, max_row=len(daily))
cats1 = Reference(ws_charts, min_col=1, min_row=1, max_row=len(daily))
chart1.add_data(data1, titles_from_data=False)
chart1.set_categories(cats1)
chart1.shape = 4
chart1.width = 18
chart1.height = 10

ws_charts.add_chart(chart1, "E1")

# LINE CHART - Gross Sales
chart2 = LineChart()
chart2.style = 10
chart2.title = "Daily Gross Sales"
chart2.y_axis.title = "Sales ($)"
chart2.x_axis.title = "Date"

data2 = Reference(ws_charts, min_col=3, min_row=1, max_row=len(daily))
chart2.add_data(data2, titles_from_data=False)
chart2.set_categories(cats1)
chart2.width = 18
chart2.height = 10

ws_charts.add_chart(chart2, "E22")

# Hide the data columns used for charts
ws_charts.column_dimensions['A'].hidden = True
ws_charts.column_dimensions['B'].hidden = True
ws_charts.column_dimensions['C'].hidden = True

print("✓ Created: Charts sheet")

# ============================================
# SAVE WORKBOOK
# ============================================
filename = 'november_2025_report.xlsx'
wb.save(filename)
print(f"\n{'='*50}")
print(f"✅ Excel report saved: {filename}")
print(f"{'='*50}")
print("\nSheets created:")
print("  1. Executive Summary - KPIs overview")
print("  2. Daily Data - Full daily breakdown with formatting")
print("  3. Customer Analysis - New vs Returning comparison")
print("  4. Charts - Visual representations")
